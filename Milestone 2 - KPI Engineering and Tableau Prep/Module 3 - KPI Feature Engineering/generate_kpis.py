"""
generate_kpis.py
Module 3: KPI Feature Engineering

Author: Yaswanth

Reads military_cleaned.csv (output of Module 2) and produces military_final.xlsx
containing two sheets:
    - Wide : one row per country, all original metrics + KPIs + metadata
    - Long : melted (country, metric, value) format for flexible Tableau use

KPI definitions used (see README / project chat for rationale):
    1. Power Index Rank Gap = GDP_rank - Power_Index_rank
         - Power Index is ranked ascending (rank 1 = strongest, lowest power_index score,
           matching GlobalFirepower's own convention).
         - GDP is ranked descending (rank 1 = largest economy).
         - Positive gap  -> country's military rank is BETTER than its economy alone would
                             predict (military "overperformer" relative to GDP).
         - Negative gap  -> country's military rank is WORSE than its GDP would predict
                             (military "underperformer" relative to GDP).
    2. Assets per Capita = (active_personnel + tanks + total_military_aircraft +
                             total_naval_fleet + total_military_helicopters) / total_population
    3. Budget-to-GDP Ratio = defense_budget_usd / gdp_usd
         - Left as NaN where gdp_usd is 0 (see gdp_is_missing flag from Module 2) since the
           ratio is undefined for those countries; avoids a fake divide-by-zero value entering
           the dashboards.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT_CSV = "military_cleaned.csv"
OUTPUT_XLSX = "military_final.xlsx"

FONT_NAME = "Arial"

# ---------------------------------------------------------------------------
# 1. Load
# ---------------------------------------------------------------------------
df = pd.read_csv(INPUT_CSV)

required_cols = [
    "country", "power_index", "gdp_usd", "defense_budget_usd", "total_population",
    "active_personnel", "tanks", "total_military_aircraft", "total_naval_fleet",
    "total_military_helicopters", "continent", "region", "alliance",
]
missing = [c for c in required_cols if c not in df.columns]
if missing:
    raise ValueError(f"Input CSV is missing required columns: {missing}")

# ---------------------------------------------------------------------------
# 2. KPI 1 - Power Index Rank Gap
#    gdp_usd is 0 for 27 countries where Module 2 flagged GDP as genuinely missing
#    (gdp_is_missing == True). Those zeros are placeholders, not real GDP values, so
#    they must NOT compete in the GDP ranking -- ranking them at the bottom would make
#    e.g. Russia/South Korea look like huge over-performers purely due to missing data.
#    We rank only countries with real GDP data and leave gdp_rank / the gap as NaN
#    for the rest.
# ---------------------------------------------------------------------------
df["power_index_rank"] = df["power_index"].rank(method="min", ascending=True).astype(int)

gdp_for_rank = df["gdp_usd"].where(~df["gdp_is_missing"])
df["gdp_rank"] = gdp_for_rank.rank(method="min", ascending=False)
df["power_index_rank_gap"] = df["gdp_rank"] - df["power_index_rank"]

# keep gdp_rank as nullable integer where present, NaN where GDP was missing
df["gdp_rank"] = df["gdp_rank"].astype("Int64")
df["power_index_rank_gap"] = df["power_index_rank_gap"].astype("Int64")

# ---------------------------------------------------------------------------
# 3. KPI 2 - Assets per Capita
# ---------------------------------------------------------------------------
df["total_assets"] = (
    df["active_personnel"]
    + df["tanks"]
    + df["total_military_aircraft"]
    + df["total_naval_fleet"]
    + df["total_military_helicopters"]
)
df["assets_per_capita"] = df["total_assets"] / df["total_population"]

# ---------------------------------------------------------------------------
# 4. KPI 3 - Budget-to-GDP Ratio
# ---------------------------------------------------------------------------
df["budget_to_gdp_ratio"] = df["defense_budget_usd"] / df["gdp_usd"].replace(0, pd.NA)

# ---------------------------------------------------------------------------
# 5. Metadata enrichment
# ---------------------------------------------------------------------------
df["is_nato"] = df["alliance"].str.upper().eq("NATO")

# sanity assertions (fail loudly rather than ship a bad file)
kpi_cols = ["power_index_rank_gap", "assets_per_capita", "budget_to_gdp_ratio"]
for col in kpi_cols:
    assert col in df.columns, f"KPI column {col} missing"
assert df["power_index_rank"].between(1, len(df)).all()
assert df["assets_per_capita"].ge(0).all()

# ---------------------------------------------------------------------------
# 6. Wide format
# ---------------------------------------------------------------------------
front_cols = [
    "country", "continent", "region", "alliance", "is_nato",
    "power_index", "power_index_rank", "gdp_rank", "power_index_rank_gap",
    "assets_per_capita", "budget_to_gdp_ratio",
]
other_cols = [c for c in df.columns if c not in front_cols]
wide_df = df[front_cols + other_cols].sort_values("power_index_rank").reset_index(drop=True)

# ---------------------------------------------------------------------------
# 7. Long format
#    One row per country per numeric metric (includes original metrics + KPIs).
#    Metadata columns are repeated on every row so Tableau needs zero joins/pivots.
# ---------------------------------------------------------------------------
id_cols = ["country", "continent", "region", "alliance", "is_nato"]
exclude_from_melt = set(id_cols) | {"gdp_is_missing"}
metric_cols = [c for c in wide_df.columns if c not in exclude_from_melt]

long_df = wide_df.melt(
    id_vars=id_cols,
    value_vars=metric_cols,
    var_name="metric",
    value_name="value",
)
long_df = long_df.sort_values(["country", "metric"]).reset_index(drop=True)

# ---------------------------------------------------------------------------
# 8. Write to Excel
# ---------------------------------------------------------------------------
with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
    wide_df.to_excel(writer, sheet_name="Wide", index=False)
    long_df.to_excel(writer, sheet_name="Long", index=False)

# ---------------------------------------------------------------------------
# 9. Formatting pass (professional font, bold header, autofit, freeze header row)
# ---------------------------------------------------------------------------
wb = load_workbook(OUTPUT_XLSX)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
body_font = Font(name=FONT_NAME)

for sheet_name in ("Wide", "Long"):
    ws = wb[sheet_name]
    ws.freeze_panes = "A2"

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    max_col = ws.max_column
    max_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.font = body_font

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        header_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        sample_lens = [
            len(str(ws.cell(row=r, column=col_idx).value))
            for r in range(2, min(max_row, 50) + 1)
        ]
        width = max([header_len] + sample_lens) + 3
        ws.column_dimensions[letter].width = min(width, 40)

wb.save(OUTPUT_XLSX)

print(f"Wrote {OUTPUT_XLSX}")
print(f"Wide sheet: {wide_df.shape[0]} rows x {wide_df.shape[1]} cols")
print(f"Long sheet: {long_df.shape[0]} rows x {long_df.shape[1]} cols")
print("\nKPI sample (top 5 by Power Index):")
print(
    wide_df[
        ["country", "power_index_rank", "gdp_rank", "power_index_rank_gap",
         "assets_per_capita", "budget_to_gdp_ratio", "is_nato"]
    ].head(5).to_string(index=False)
)
